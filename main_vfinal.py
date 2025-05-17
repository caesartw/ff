
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
from collections import defaultdict
import openpyxl
import random
import os
import pandas as pd
from dataclasses import dataclass, field

@dataclass
class ProjectData:
    name: str
    required_count: int = 1
    note: str = ""
    cells: list = field(default_factory=list)

class AssignmentApp:
    def __init__(self, master):
        self.master = master
        self.master.title("地點導向工作分配系統")
        self.staff = []
        self.projects_by_location = defaultdict(list)
        self.locations = []
        self.assignments = defaultdict(list)
        self.assignment_widgets = {}
        self.file_paths = {}
        self.assignment_grids = {}
        self.setup_ui()

    def setup_ui(self):
        frame = ttk.Frame(self.master)
        frame.pack(padx=10, pady=10, fill='x')
        ttk.Button(frame, text="載入人員", command=self.load_staff).grid(row=0, column=0, padx=5)
        ttk.Button(frame, text="載入項目", command=self.load_projects).grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="載入地點", command=self.load_locations).grid(row=0, column=2, padx=5)
        ttk.Button(frame, text="開始分配", command=self.start_assignment).grid(row=1, column=0, pady=10)
        ttk.Button(frame, text="匯出結果", command=self.export_results).grid(row=1, column=1, pady=10)
        ttk.Button(frame, text="測試自動分配", command=self.test_auto_assign).grid(row=1, column=2, pady=10)
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(expand=True, fill='both')

    def load_staff(self):
        path = filedialog.askopenfilename(title="選擇人員.csv")
        if not path: return
        with open(path, encoding="utf-8") as f:
            self.staff = [line.strip() for line in f if line.strip()]
        messagebox.showinfo("完成", f"載入 {len(self.staff)} 位人員")

    def load_projects(self):
        path = filedialog.askopenfilename(title="選擇項目.csv")
        if not path: return
        self.file_paths['projects'] = path
        self.projects_by_location.clear()
        with open(path, encoding="utf-8") as f:
            reader = csv.reader(f)
            uid_counter = defaultdict(int)
            for row in reader:
                if len(row) < 2:
                    continue
                loc = row[0].strip()
                name = row[1].strip()
                note = row[2].strip() if len(row) >= 3 else ""
                cells = [c.strip() for c in row[3].split("|")] if len(row) >= 4 and row[3].strip() else []
                count = int(row[4]) if len(row) >= 5 and row[4].isdigit() else 1
                uid_counter[(loc, name)] += 1
                unique_name = f"{name}#{uid_counter[(loc, name)]}"
                self.projects_by_location[loc].append(ProjectData(unique_name, count, note, cells))
        messagebox.showinfo("完成", "已載入項目")

    def load_locations(self):
        path = filedialog.askopenfilename(title="選擇地點.csv")
        if not path: return
        with open(path, encoding="utf-8") as f:
            self.locations = [line.strip() for line in f if line.strip()]
        messagebox.showinfo("完成", f"載入 {len(self.locations)} 個地點")

    def start_assignment(self):
        for tab in self.notebook.tabs():
            self.notebook.forget(tab)
        self.assignment_widgets.clear()
        self.assignments.clear()
        self.assignment_grids = {}
        for loc in self.locations:
            canvas = tk.Canvas(self.notebook)
            frame = ttk.Frame(canvas)
            vsb = ttk.Scrollbar(canvas, orient='vertical', command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            canvas.pack(side='left', fill='both', expand=True)
            vsb.pack(side='right', fill='y')
            canvas.create_window((0, 0), window=frame, anchor='nw')
            frame.bind('<Configure>', lambda e, c=canvas: c.configure(scrollregion=c.bbox('all')))
            canvas.bind('<MouseWheel>', lambda e, c=canvas: c.yview_scroll(int(-1*(e.delta/120)), 'units'))
            self.notebook.add(canvas, text=loc)
            self.assignment_widgets[loc] = []
            projects = self.projects_by_location.get(loc, [])
            rows, cols = 10, 6
            grid = ttk.Frame(frame)
            row_offset = 0
            grid.pack(fill='both', expand=True)
            self.assignment_grids[loc] = grid
            for idx, project in enumerate(projects):
                for p in range(project.required_count):
                    r, c = divmod(row_offset, cols)
                    row_offset += 1
                    cell = ttk.Frame(grid, padding=2)
                    cell.grid(row=r, column=c, sticky='w')
                    display_name = project.name.split('#')[0]
                    display_text = f"{display_name}（{project.note}）" if project.note else display_name
                    ttk.Label(cell, text=display_text, wraplength=120, justify='left').pack(anchor='w')
                    var = tk.StringVar()
                    cb = ttk.Combobox(cell, textvariable=var, values=self.get_available_staff(loc), state='readonly')
                    cb.pack(fill='x')
                    cb.bind('<<ComboboxSelected>>', lambda e, l=loc: self.update_assignments(l))
                    ttk.Button(cell, text='清除', command=lambda v=var, l=loc: self.clear_assignment(v, l)).pack(pady=1)
                    self.assignment_widgets[loc].append((project, var, cb))

    def get_available_staff(self, loc):
        used = {s for l in self.assignments if l != loc for _, s in self.assignments[l]}
        return [s for s in self.staff if s not in used]

    def clear_assignment(self, var, loc):
        var.set('')
        self.update_assignments(loc)

    def update_assignments(self, loc):
        self.assignments[loc] = []
        for proj, var, _ in self.assignment_widgets[loc]:
            if var.get():
                self.assignments[loc].append((proj, var.get()))
        for proj, var, cb in self.assignment_widgets[loc]:
            vals = self.get_available_staff(loc)
            cb['values'] = vals
            if var.get() not in vals:
                var.set('')

    def export_weight_table(self):
        counter = defaultdict(lambda: defaultdict(int))
        for loc, entries in self.assignments.items():
            for proj, staff in entries:
                counter[staff][loc] += 1

        weight_file = "人員權重.csv"
        if os.path.exists(weight_file):
            df_old = pd.read_csv(weight_file, index_col=0)
        else:
            df_old = pd.DataFrame()

        df_new = pd.DataFrame(counter).T.fillna(0).astype(int)
        combined = df_old.add(df_new, fill_value=0).fillna(0).astype(int)
        combined["總計"] = combined[self.locations].sum(axis=1)

        try:
            combined.to_csv(weight_file, encoding='utf-8-sig')
        except PermissionError:
            messagebox.showerror("儲存錯誤", "請先關閉『人員權重.csv』再試一次。")

    def test_auto_assign(self):
        answer = messagebox.askyesno("自動分配模式", "是否要依據『人員權重.csv』進行分配？")
        if answer:
            path = filedialog.askopenfilename(title="選擇人員權重.csv", filetypes=[("CSV 檔案", "*.csv")])
            if not path:
                messagebox.showwarning("取消", "未選擇權重檔案，使用原始方式分配")
                return self.random_assign()
            try:
                df = pd.read_csv(path, index_col=0)
                for loc in self.assignment_widgets:
                    if loc not in df.columns:
                        continue
                    ranked = df[loc].sort_values(ascending=False)
                    pool = list(ranked.index)
                    ptr = 0
                    for proj, var, _ in self.assignment_widgets[loc]:
                        while ptr < len(pool):
                            candidate = pool[ptr]
                            ptr += 1
                            if candidate in self.get_available_staff(loc):
                                var.set(candidate)
                                break
                    self.update_assignments(loc)
            except Exception as e:
                messagebox.showerror("錯誤", f"載入權重檔案失敗：{e}")
        else:
            self.random_assign()

    def random_assign(self):
        if not self.staff or not self.assignment_widgets:
            messagebox.showerror('錯誤','請先載入人員與項目並開始分配')
            return
        ppl = self.staff[:]
        random.shuffle(ppl)
        locs = list(self.assignment_widgets.keys())
        per = len(ppl)//len(locs)
        rem = len(ppl)%len(locs)
        i = 0
        for idx, l in enumerate(locs):
            cnt = per + (1 if idx < rem else 0)
            pool = ppl[i:i+cnt]
            i += cnt
            for proj, var, _ in self.assignment_widgets[l]:
                if not var.get() and pool:
                    var.set(pool.pop(0))
            filled = [var.get() for _, var, _ in self.assignment_widgets[l] if var.get()]
            for proj, var, _ in self.assignment_widgets[l]:
                if not var.get() and filled:
                    var.set(random.choice(filled))
            self.update_assignments(l)

    def export_results(self):
        try:
            path = filedialog.askopenfilename(title='選擇 Excel 模板', filetypes=[('Excel', '*.xlsx')])
            if not path:
                return
            from shutil import copyfile
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            out_path = path.replace('.xlsx', f'_分配結果_{timestamp}.xlsx')
            copyfile(path, out_path)
            wb = openpyxl.load_workbook(out_path)
            ws = wb.active
            map_path = self.file_paths.get('projects') or filedialog.askopenfilename(title='選擇 項目對應 CSV', filetypes=[('CSV','*.csv')])
            if not map_path:
                return
            df_map = pd.read_csv(map_path)
            grouped = defaultdict(lambda: defaultdict(list))
            for loc, entries in self.assignments.items():
                for proj, staff in entries:
                    proj_base = proj.name.split('#')[0]
                    rows = df_map[(df_map['地點'] == loc) & (df_map['項目名稱'] == proj_base)]
                    for _, r in rows.iterrows():
                        for c in str(r['儲存格']).split('|'):
                            cell = c.strip()
                            if not cell:
                                continue
                            grouped[cell][proj_base].append((loc, staff))
            content = defaultdict(list)
            for cell, proj_map in grouped.items():
                for proj_name, entries in proj_map.items():
                    if cell == 'D31':
                        loc_groups = defaultdict(list)
                        for loc, staff in entries:
                            loc_groups[loc].append(staff)
                        for loc, staff_list in loc_groups.items():
                            line = f"【{loc}】{proj_name}：" + " ".join(staff_list)
                            content[cell].append(line)
                    else:
                        staff_list = [s for _, s in entries]
                        line = f"{proj_name}：" + " ".join(staff_list)
                        content[cell].append(line)
            for cell, lines in content.items():
                ws[cell] = "\n".join(lines)
                ws[cell].alignment = openpyxl.styles.Alignment(wrap_text=True)
                ws.row_dimensions[ws[cell].row].height = 15 * len(lines)
            wb.save(out_path)
            self.export_weight_table()
            messagebox.showinfo("完成", f"已儲存：{out_path}")
        except Exception as e:
            messagebox.showerror("匯出錯誤", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = AssignmentApp(root)
    root.mainloop()
